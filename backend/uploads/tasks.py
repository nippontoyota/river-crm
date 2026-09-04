import csv
import io
import re
from datetime import datetime

from celery import shared_task
from django.db import transaction
from django.utils import timezone

from leads.models import Lead
from leads.serializers import configured_source, configured_values
from .models import UploadBatch, UploadRow
from .storage import download_bytes

def value(row, *names):
    lowered = {str(key).strip().lower(): value for key, value in row.items() if key}
    return next((str(lowered[name]).strip() for name in names if lowered.get(name) not in (None, "")), "")


def normalize_phone(phone):
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    return digits if len(digits) == 10 else ""


def classify_source(raw):
    source = configured_source(raw)
    return source or raw.strip(), raw, "" if source else "Choose a lead source from Admin Lists."


def parse_date(raw):
    for format_string in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, format_string).date()
        except ValueError:
            continue
    return timezone.localdate()


def invalid_model_error(model):
    allowed = configured_values("models")
    return "Choose a vehicle model from Admin Lists." if model and allowed and model not in allowed else ""


def read_rows(filename, content):
    if filename.lower().endswith(".csv"):
        yield from csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        return
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, values))


@shared_task
def parse_upload_batch(batch_id):
    batch = UploadBatch.objects.get(id=batch_id)
    try:
        content = download_bytes(batch.storage_path)
        parsed_rows = []
        skipped = 0
        for row_number, row in enumerate(read_rows(batch.filename, content), start=2):
            name = value(row, "name", "customer name")
            phone = normalize_phone(value(row, "phone", "mobile"))
            if not name and not phone:
                continue
            error = "" if name and phone else "Name and phone are required."
            enquiry_date = parse_date(value(row, "date", "enquiry date"))
            if enquiry_date > timezone.localdate():
                error = "Enquiry date cannot be in the future."
            source, source_label, source_error = classify_source(value(row, "source"))
            model_interest = value(row, "model", "vehicle interest", "model / vehicle interest")
            error = error or source_error or invalid_model_error(model_interest)
            if error:
                skipped += 1
            parsed_rows.append({"row_number": row_number, "phone": phone, "validation_error": error, "data": {"name": name, "email": value(row, "email"), "source": source, "source_label": source_label, "campaign": value(row, "campaign"), "model_interest": model_interest, "city": value(row, "city", "location"), "enquiry_date": enquiry_date.isoformat()}})
        existing_leads = {}
        for lead in Lead.objects.filter(phone__in={row["phone"] for row in parsed_rows if row["phone"]}, deleted_at__isnull=True).only("id", "phone").order_by("id"):
            existing_leads.setdefault(lead.phone, lead)
        staged = []
        first_uploaded_phone = {}
        duplicates = 0
        for row in parsed_rows:
            data = row["data"].copy()
            duplicate = existing_leads.get(row["phone"]) if not row["validation_error"] else None
            resolution = UploadRow.Resolution.IMPORT
            if duplicate:
                data["_duplicate_type"] = "CRM"
                resolution = UploadRow.Resolution.SKIP
                duplicates += 1
            elif not row["validation_error"] and row["phone"] in first_uploaded_phone:
                first_row = first_uploaded_phone[row["phone"]]
                data["_duplicate_type"] = "FILE"
                data["_duplicate_label"] = f"Row {first_row['row_number']} - {first_row['data'].get('name') or 'first matching row'}"
                resolution = UploadRow.Resolution.SKIP
                duplicates += 1
            elif not row["validation_error"] and row["phone"]:
                first_uploaded_phone[row["phone"]] = row
            staged.append(UploadRow(batch=batch, row_number=row["row_number"], normalized_phone=row["phone"], validation_error=row["validation_error"], duplicate_of=duplicate, resolution=resolution, data=data))
        with transaction.atomic():
            UploadRow.objects.filter(batch=batch).delete()
            UploadRow.objects.bulk_create(staged)
            batch.total_rows = len(staged)
            batch.parsed_ok = len([row for row in staged if not row.validation_error and row.resolution != UploadRow.Resolution.SKIP])
            batch.duplicates_found = 0
            batch.skipped = skipped + duplicates
            batch.status = UploadBatch.Status.READY
            batch.save(update_fields=["total_rows", "parsed_ok", "duplicates_found", "skipped", "status"])
    except Exception as error:
        batch.status = UploadBatch.Status.FAILED
        batch.error_message = str(error)[:1000]
        batch.save(update_fields=["status", "error_message"])
        raise
